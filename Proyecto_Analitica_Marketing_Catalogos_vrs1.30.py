# -*- coding: utf-8 -*-
"""
Created on Tue Feb 20 10:50:41 2018

@author: jtaquia
"""

from __future__ import print_function
import csv
import os
import glob
import io
import sys
from tkinter.colorchooser import askcolor
from contextlib import redirect_stdout
from tkinter import Tk
from tkinter import ttk
import tkinter as tk
from tkinter import messagebox
import subprocess
from tkinter import filedialog
from tkinter import *
import numpy as np
import argparse
import cv2

def fnPDF_FindText(xFile, xString):
    
        import PyPDF2
        #import subprocess
        import os
       # print(xFile)
        #print(xString)
        # xfile : the PDF file in which to look
        # xString : the string to look for
        import re
        PageFound = -1
        pdfDoc = PyPDF2.PdfFileReader(open(xFile, "rb"))
        for i in range(0, pdfDoc.getNumPages()):
            
            #print(pdfDoc.getNumPages())
            content = ""
            content += pdfDoc.getPage(i).extractText() + "\n"
            #print(type(content))
            #content1 = content.encode('ascii', 'ignore').lower()
            
            content1 = content.encode('utf-8') #, 'ignore').lower()
            #content1 = content.encode('ascii', 'ignore').lower()
            ResSearch = re.search(xString, content1)
            if ResSearch is not None:
               PageFound = i
    
               print('el valor es ' + str(xFile))
              
               print(' el resultado SI esta en la página: '+ str(PageFound))
               #break
              
               #gge= os.getcwd()                       
               #print(gge)
               #path_to_pdf = os.path.abspath( gge + xFile )
               #path_to_acrobat = os.path.abspath('C://Program Files (x86)//Adobe//Acrobat Reader DC//Reader//AcroRd32.exe') 
               #pdfFileObj = open('D://Documentos//JOSE ANTONIO//COMPUTER VISION//LECTURA DE PDF EN PYTHON//INPUT//2018-0 SIL ANALISIS PREDICTIVO.pdf', 'rb')
               #var = PageFound
               #process = subprocess.Popen([path_to_acrobat, '/A', 'page=' + str(var), path_to_pdf], shell=False, stdout=subprocess.PIPE)
               #process.wait()
                
            
            else:
               
                print('en el archivo no esta la palabra')
         
        
        return PageFound
    

class ClassA(object):
    def __init__(self):
        currdir = os.getcwd()
        self.var1 = filedialog.askdirectory(parent=root, initialdir=currdir, title='Favor seleccione un directorio')   
        #self.var2 = filedialog.askdirectory(parent=root, initialdir=currdir, title='Favor seleccion donde guardará el resultado')   
        #print(self.var1)
        
    def abrir1(tempdir,tempdir2):
       
       #print(tempdir, tempdir2) 
        
       #root = Tk()
       #root.withdraw() #use to hide tkinter window
       currdir = os.getcwd()
       tempdir = filedialog.askdirectory(parent=root, initialdir=currdir, title='Favor seleccione un directorio')   
       #filename = filedialog.askopenfile(parent=root, initialdir=currdir, title='Please select a directory')
       #print(filename.name)
    
       
       
       if len(tempdir) > 0:
          
          messagebox.showinfo("Information","Usted eligió la carpeta  %s" % tempdir + " donde estan los pdf a analizar")
           #print ("Usted eligió %s" % tempdir)
       else:
           messagebox.showwarning("Warning","Ingrese una ruta")
    
    #def abrir2():
       #root = Tk() 
       #root.withdraw() #use to hide tkinter window
       messagebox.showinfo("Information","Seleccione la carpeta donde se guardará los resultados")
       
       currdir = os.getcwd()
       tempdir2 = filedialog.askdirectory(parent=root, initialdir=currdir, title='Favor seleccion donde guardará el resultado')   
       #filename = filedialog.askopenfile(parent=root, initialdir=currdir, title='Please select a directory')
       #print(filename.name)
      
       
       if len(tempdir2) > 0:
           
           messagebox.showinfo("Information","Usted eligió %s" % tempdir2)
           #print ("Usted eligió %s" % tempdir2)
        
       
       pdfDir = tempdir.replace('/','//')+ "//" 
       txtDir = tempdir2.replace('/','//')+ "//" 
       #action_with_arg = partial(iteracion, txtDir)
       print(pdfDir, txtDir)     
          
       convertMultiple(pdfDir, txtDir)
    
def convertMultiple(pdfDir, txtDir):
   
    import os
    #fdir = pdfDir
    if pdfDir == "": pdfDir = os.getcwd() + "\\" #if no pdfDir passed in 
    for pdf in os.listdir(pdfDir): #iterate through pdfs in pdf directory
            fileExtension = pdf.split(".")[-1]
            if fileExtension == "pdf":
                pdfFilename = pdfDir + pdf 
                #print(pdfFilename)
                text = layout(open(pdfFilename,'rb'),txtDir,pdf) #get string of text content of pdf
    
    messagebox.showwarning("Warning","Se culminó con la extracción de texto de los pdf colocados en la carpeta : " + pdfDir + " y los resultados estan en : " + txtDir)

def layout(pdf_file, txtDir2,pdf2):
    
    from pdfminer.pdfparser import PDFParser, PDFDocument
    from pdfminer.layout import LAParams
    from pdfminer.converter import PDFPageAggregator
    from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
    from pdfminer.layout import LTTextBox, LTTextLine
    
    parser = PDFParser(pdf_file)
    doc = PDFDocument(pdf_file)
    parser.set_document(doc)
    doc.set_parser(parser)
    doc.initialize('')
    
    #Create resource manager
    rsrcmgr = PDFResourceManager()
    # Set parameters for analysis.
    laparams = LAParams()
    # Create a PDF page aggregator object.
    device = PDFPageAggregator(rsrcmgr, laparams=laparams)
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    
    textFilename = txtDir2 + pdf2 + ".txt"
    text_file = open(textFilename, "w")
    
    #textFile = open(textFilename, "w") #make text file
 
    
    for page in doc.get_pages():
        interpreter.process_page(page)
        print(page.pageid)
        print(page.mediabox)
        # receive the LTPage object for the page.
        layout = device.get_result()
    
        for element in layout:
                 if isinstance(element, LTTextBox) or isinstance(element, LTTextLine):
                    text =element.get_text().encode('utf-8')
                    #print(text)
                    text_file.write("%s" % text)


    messagebox.showwarning("Warning","Se extrajo el texto en" + textFilename)


def hello():
    print ("hello!")

def salida():
    root.quit()
    root.destroy()

class myDict(dict):

    def __init__(self):
        self = dict()

    def add(self, key, value):
        self[key] = value
    
class GUI3():
    
    def __init__(self,root3):
        self.label3 = tk.Label(root3,text="Analizador de imágenes vrs 1. Diseño y programación: José Antonio Taquía Gutiérrez", width = 120, height=4, fg = "blue", font=("Helvetica",10,"bold"))
        self.label3.place(x=3, y  = 2)                    
 
class GUI2():
    
    
   # class Example(object):
   
        #self.itsProblem = "problem"
    
    
    
    def __init__(self,root2):
                
            self.root2 = root2 # root2 is a passed Tk object
            #self.root4 = root4
            self.Entry = tk.Entry(root2, width = 30)
            self.Entry.place(x=250, y = 40)
            self.frame = tk.Frame(self.root2)#grid(row=6, column= 6)
            fd = ('D://Documentos//JOSE ANTONIO//COMPUTER VISION//SQLITE//ejemplo.csv')            
            with open(fd,'r') as csv_file:
                csv_reader = csv.reader(csv_file)
                Passy=[]
                for line in csv_reader:
                    Passy.append(line)
            print(Passy)  
            #self.cbox_value = StringVar()
            self.cbox = ttk.Combobox(self.root2,width=30,height=10, values = sorted(Passy)) #,textvariable=self.cbox_value) 
            self.cbox.place(x=250,y=80)
            #self.locationBox = Combobox(self.master, textvariable=self.box_value)
            self.cbox.bind("<<ComboboxSelected>>", self.justamethod)
            #self.locationBox['values'] = ('one', 'two', 'three')
            #self.locationBox.current(0)
            #self.Entry.grid(row=2, column=1, sticky=W)
            self.canvas = tk.Canvas()
            self.label2 = tk.Label(self.root2, text="Ingrese búsqueda:", width = 20, height=4, font=("Helvetica", 12))
            self.label2.place(x =40 , y =20)
            self.button2 = tk.Button(self.root2, text='Empezar', command= self.printing1, width = 10, height=2, font=("Helvetica", 12))
            self.button2.place(x =500 , y =20)
            self.button3 =  tk.Button(self.root2, text='Reiniciar', command= self.Reinicio, width = 10, height=2, font=("Helvetica", 12))
            self.button3.place(x =500 , y =80)
            self.button4 =  tk.Button(self.root2, text='Mostrar', command= self.mostrar500, width = 10, height=2, font=("Helvetica", 12))
            self.button4.place(x =500 , y =140)
            self.button5 =  tk.Button(self.root2, text='Pixels', command= self.pixels, width = 10, height=2, font=("Helvetica", 12))
            self.button5.place(x =500 , y =200)
            self.button6 =  tk.Button(self.root2, text='Precios', command= self.prueba111, width = 10, height=2, font=("Helvetica", 12))
            self.button6.place(x =500 , y =260)
            self.button7 =  tk.Button(self.root2, text='Salir', command= self.removethis2, width = 10, height=2, font=("Helvetica", 12))
            self.button7.place(x =500 , y =320)
            color = '#fee2a1'
            self.Text = tk.Text(root2, height=10, width=45, bg = color)
            self.Text.place(x =40 , y =120)
            self.lstbox = tk.Listbox(self.root2,height=10, width=60,bg = color,selectmode=EXTENDED)
            self.lstbox.place(x =40 , y =320)
            #self.lstbox.pack(fill=BOTH, expand=1)
            self.w = self.lstbox #.widgetName(Listbox(self.root2))
            self.index = int(self.w.curselection()[0])
            self.value = self.w.get(self.index)
                
    
    def prueba111(self):
            
            import re
            import os
            import PyPDF2
            import xlwt
            import xlsxwriter
            from pandas import DataFrame
            from tempfile import TemporaryFile
            import openpyxl
            from openpyxl import Workbook
         
            
            
            
           # workbook = xlsxwriter.Workbook('demo5.xlsx')
           # worksheet = workbook.add_worksheet()
            #print(self.myd)
            #print(type(self.myd))
            
            
#==============================================================================
#              En printing1            
#             part1 = 'En el archivo '+ j
#            #part2 = 'En la página : '+ str(var101)
#            part2 = 'En la página '+ " ".join([str(x)  for x in Lista_pickin] ) #+ str(len(Lista_pickin))
#            #print(part2)
#            #self.myd.add(part1 , part2)
# 
#==============================================================================
            
            
            dictlist = []
            
            wb = openpyxl.Workbook()
            wb.create_sheet("INICIO")
            
            for key, value in self.myd.items():
                temp = [key,value]
                dictlist.append(temp)
            
             
            #print(dictlist)                        
            #print(len(dictlist))
            #print(type(dictlist))
            #print(dictlist[0])
            #print(len(dictlist[0]))
            #print(dictlist[1])
            #print(len(dictlist[1]))
            
            
            
            for m in range(len(dictlist)):
                str1 = ''.join(dictlist[m])
                #print(str1)
                juanji = re.split('\.+|\s+', str1)
                #print(juanji)
                
                pais = juanji[3]
                campanha =juanji[4]
                paginas = juanji[8:]
                size_total=[]
                size_total.append(pais + " " + campanha)
                #print(pais)
                #print(campanha)
                #print(paginas)
                
                
                
                for j in range(len(paginas)):
                                
                    Passy_X = paginas[j]
                    #print('la pagina buscada es la numero' + Passy_X)
                    currdir = os.getcwd()
                    #print(currdir)
                    #tempdir = tk.filedialog.askdirectory( initialdir= currdir, title='Favor seleccion donde guardará el resultado')   
                    #os.chdir(tempdir)
                    #os.chdir('D://Yanbal_DLO//PROYECTO COMPUTER VISION//CATALOGOS PERU//')
                    #k= glob.glob('*.pdf')
                    
                    #print(k)
                    
                    
                    
                    #for i in k:
                   
                    #txtDir0902 = tempdir.replace('/','//')+ "//" + i 
                    #txtDir0902 = currdi#r.replace('/','//')+ "//" + pais + " "+ campanha + ".pdf" 
                    txtDir0902 = currdir + "/" + pais + " "+ campanha + ".pdf" 
                            
                    #print(txtDir0902)
                    #nombre = re.split('\.+',i)
                    #size_total.append(nombre[0])
                    Lista_pickin = []
                    PageFound = -1
                    pdfDoc = PyPDF2.PdfFileReader(open(txtDir0902,'rb'))
                    counter =0
                    content = ""
                    content += pdfDoc.getPage(int(Passy_X)).extractText() #+ "\n"
                    content1 = content.encode('utf-8') #, 'ignore').lower()
                    #print(content)
                    juanjose = re.split('\s+|\.+',content)
                    #juanjose = re.split('\s',content)
                    #juanjose = re.split('\s',content)
                    #juanjose = re.split('\s+|\.+|\w+',content)
                    #juanjose = re.split('\w+',content)
                    #juanjose = re.split('\.*',content)
                    #juanjose = re.split('\n',content)
                    
                    #mylist = ['abc123', 'def456', 'ghi789', 'ABC987', 'aBc654']
                    #sub = 'abc'
                    #print ("\n".join(s for s in mylist if sub.lower() in s.lower())) # ENCUENTRA LA BUSQUEDA CASE INSENTITIVITY
                    
                    
                    target = 'Precio'
                    matching = [s for s in juanjose if target in s] # ENCUENTRA EL TARGET EN UN ELEMENTO DE UNA LISTA
                    #print(matching)
                    #print(len(juanjose))
                    def linear_search(data, target,Mari,Juanpi):
                        Juanpi=Juanjito
                        Mari =Kari
                        #print(len(data))
                        for i in range(len(data)):
                           # matching = [s for s in data if target in s] # ENCUENTRA EL TARGET EN UN ELEMENTO DE UNA LISTA
                           # print(matching)
                          if target in data[i]:
                             Juanpi.append(i)
                             Mari.append(i+1)
                             #print(data[i])
                               #print(i)
                               #print(Mari)
                        return Juanpi, Mari
                            
                        
                    Juanjito=[]
                    Kari=[]
                    linear_search(juanjose,target, Kari,Juanjito)    
                    
                    Juanpablo =[]
                    Mayita = []
                    
                    for i in range(len(Kari)):
                        Juanpablo.append(juanjose[Juanjito[i]])
                        Mayita.append(juanjose[Kari[i]])
                    
                    mylict2 = (list(zip(Juanpablo,Mayita)))
                    
                    #print(mylict2)
                    
                    target = 'OFERTA'    
                    Juanjito=[]
                    Kari=[]
                    linear_search(juanjose,target, Kari,Juanjito)    
                    
                    Juanpablo =[]
                    Mayita = []
                    
                    for i in range(len(Kari)):
                        
                        Juanpablo.append(juanjose[Juanjito[i]])
                        Mayita.append(juanjose[Kari[i]])
                    
                    mylict3 = (list(zip(Juanpablo,Mayita)))
                    #print(mylict3)
                    
                    mylict4 = list(zip(mylict2,mylict3))
                    #print(mylict4)
                    
                    target = '%'    
                    matching = [s for s in juanjose if target in s] # ENCUENTRA EL TARGET EN UN ELEMENTO DE UNA LISTA
                    #print(matching)
                    def linear_search(data, target,Juanpi,Mari):
                        Juanpi=Juanjito
                        Mari=Kari
                        #print(len(data))
                        for i in range(len(data)):
                           # matching = [s for s in data if target in s] # ENCUENTRA EL TARGET EN UN ELEMENTO DE UNA LISTA
                           # print(matching)
                          if target in data[i]:
                             Juanpi.append(i)
                             Mari.append(i+1)
                             #print(data[i])
                               #print(i)
                               #print(Mari)
                        return Juanpi, Mari
                    
                    
                    
                    Kari =[]
                    Juanjito=[]
                    linear_search(juanjose,target, Juanjito,Kari)    
                    #print(len(Juanjito))
                    #print(Juanjito)
                    #print(len(Kari))
                    #print(Juanjito)
                    
                    Juanpablo =[]
                    Mayita =[]
                    
                    for i in range(len(Juanjito)):
                    
                        Juanpablo.append(juanjose[Juanjito[i]])
                        Mayita.append(juanjose[Kari[i]])
                    
                    mylict5 = (list(zip(Juanpablo,Mayita)))
                   # print('el valor de mylict5 es:' +  " ".join([str(x)  for x in mylict5]))
                    
                    mylict6 = list(zip(size_total,mylict2,mylict3,mylict5))
                    
                    #trazos = re.split('\s+|\.+',mylict6)
                    #trazos = re.split(','," ".join([str(x )  for x in mylict6]))
                    print(mylict6)
                    #print(type(mylict6))
                    
                    if not mylict6:
                        print("List is empty")
#==============================================================================
#                     print(mylict6)
#                     print('primera ' + str(len(mylict6)))
#                     print(('el valor de mylict6 es:' +  " ".join([str(x )  for x in mylict6])))
#                     
#==============================================================================
                    else:
                        print(len(mylict6))
                        sheet2 = wb.get_sheet_by_name("INICIO")
                        row_count = sheet2.max_row
                        column_count = sheet2.max_column
                        j=sheet2.cell(row = row_count+1, column = column_count)
                        print(str(row_count))
                        j.value = " ".join([str(x )  for x in mylict6])
                        
                        sheet2.cell(row= row_count+1, column = column_count)
                
            os.chdir('D://Documentos//JOSE ANTONIO//COMPUTER VISION//PDF')
            import time
            
            timestr = time.strftime("%Y%m%d-%H%M%S")
            nombre = self.Entry.get()
            
            nombre_completo = nombre + ' ' + timestr + ".xls"
            wb.save(nombre_completo)
            messagebox.showwarning("Warning","Se culminó con exportar los precios de " + nombre + " y los resultados estan en : " + os.getcwd() + ' con el nombre : ' + nombre_completo) 
                
#==============================================================================
#                     for k in range(len(mylict6)):
#                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                 
# #==============================================================================
# #                         print(len((mylict6)))
# #                         print(mylict6[0]) 
# #                         
# #==============================================================================
#                         row_count = sheet2.max_row
#                         #print('el valor de row_count: ' + str(row_count))
#                         
#                         if k==0:
#                             
#                             j=sheet2.cell(row = k+1, column = 1)
#                             j.value = mylict6[0] 
#                             print('el valor de j.value : '+ str(j.value))
#                             sheet2.cell(row=k+1,column = 1) 
#                             
#                         else:
#                            
#                             j=sheet2.cell(row = row_count, column = k+1)
#                             j.value =mylict6[k]  
#                             sheet2.cell(row= row_count, column = 1+k)
#==============================================================================
                            
                    #for p in range(len(mylict6)):
                    #    str100 = ''.join(mylict6[p]) 
                    #    print(str100)
                    
                    
                    #worksheet.write(m, 0, str100)
    
            #print(os.getcwd())
            #print('Primera ' + str(len(mylict6)))
            
            
               
                    #df = DataFrame({'Stimulus Time': l1, 'Reaction Time': l2})
                    #df.to_excel('test.xlsx', sheet_name='sheet1', index=False)
                    
                    #supersecretdata = mylict6
                    #print(len(mylict6))
                  
                    
                    #for i,e in enumerate(supersecretdata):
                    #    sheet1.write(i,2,e)
                        
                    #name = "resultado_ofertas.xls"
                    #book.save(name)
                    #book.save(TemporaryFile())
                        
                
                #for i in len(dictlist):
                #    print(i)
                #juanji = re.split('-', dictlist[0])
                #print('los valores de juanji son:' + juanji)
                #juampi = re.split('-',dictlist[1].encode('utf-8'))
                #print('los valores de juampi son:' + juampi)
            
            
            
            
            
            
    def Analisis_Precios(self):

            import tkinter as tk
            from tkinter import filedialog
            import re
            from nltk import word_tokenize
            import os
            import PyPDF2
            import glob
                        
                    
            currdir = os.getcwd()
            tempdir = tk.filedialog.askdirectory( initialdir= currdir, title='Favor seleccion donde guardará el resultado')   
            os.chdir(tempdir)
            #os.chdir('D://Yanbal_DLO//PROYECTO COMPUTER VISION//CATALOGOS PERU//')
            k= glob.glob('*.pdf')
            
            print(k)
            
            
            
            for i in k:
                size_total=[]
                txtDir0902 = tempdir.replace('/','//')+ "//" + i 
                nombre = re.split('\.+',i)
                size_total.append(nombre[0])
                Lista_pickin = []
                PageFound = -1
                pdfDoc = PyPDF2.PdfFileReader(open(txtDir0902,'rb'))
                counter =0
                content = ""
                content += pdfDoc.getPage(88).extractText() #+ "\n"
                content1 = content.encode('utf-8') #, 'ignore').lower()
                #print(content)
                juanjose = re.split('\s+|\.+',content)
                #juanjose = re.split('\s',content)
                #juanjose = re.split('\s',content)
                #juanjose = re.split('\s+|\.+|\w+',content)
                #juanjose = re.split('\w+',content)
                #juanjose = re.split('\.*',content)
                #juanjose = re.split('\n',content)
                
                #mylist = ['abc123', 'def456', 'ghi789', 'ABC987', 'aBc654']
                #sub = 'abc'
                #print ("\n".join(s for s in mylist if sub.lower() in s.lower())) # ENCUENTRA LA BUSQUEDA CASE INSENTITIVITY
                
                
                target = 'Precio'
                matching = [s for s in juanjose if target in s] # ENCUENTRA EL TARGET EN UN ELEMENTO DE UNA LISTA
                #print(matching)
                #print(len(juanjose))
                def linear_search(data, target,Mari,Juanpi):
                    Juanpi=Juanjito
                    Mari =Kari
                    #print(len(data))
                    for i in range(len(data)):
                       # matching = [s for s in data if target in s] # ENCUENTRA EL TARGET EN UN ELEMENTO DE UNA LISTA
                       # print(matching)
                      if target in data[i]:
                         Juanpi.append(i)
                         Mari.append(i+1)
                         #print(data[i])
                           #print(i)
                           #print(Mari)
                    return Juanpi, Mari
                        
                    
                Juanjito=[]
                Kari=[]
                linear_search(juanjose,target, Kari,Juanjito)    
                
                Juanpablo =[]
                Mayita = []
                
                for i in range(len(Kari)):
                    Juanpablo.append(juanjose[Juanjito[i]])
                    Mayita.append(juanjose[Kari[i]])
                
                mylict2 = (list(zip(Juanpablo,Mayita)))
                
                #print(mylict2)
                
                target = 'OFERTA'    
                Juanjito=[]
                Kari=[]
                linear_search(juanjose,target, Kari,Juanjito)    
                
                Juanpablo =[]
                Mayita = []
                
                for i in range(len(Kari)):
                    
                    Juanpablo.append(juanjose[Juanjito[i]])
                    Mayita.append(juanjose[Kari[i]])
                
                mylict3 = (list(zip(Juanpablo,Mayita)))
                #print(mylict3)
                
                mylict4 = list(zip(mylict2,mylict3))
                #print(mylict4)
                
                target = '%'    
                matching = [s for s in juanjose if target in s] # ENCUENTRA EL TARGET EN UN ELEMENTO DE UNA LISTA
                #print(matching)
                def linear_search(data, target,Juanpi,Mari):
                    Juanpi=Juanjito
                    Mari=Kari
                    #print(len(data))
                    for i in range(len(data)):
                       # matching = [s for s in data if target in s] # ENCUENTRA EL TARGET EN UN ELEMENTO DE UNA LISTA
                       # print(matching)
                      if target in data[i]:
                         Juanpi.append(i)
                         Mari.append(i+1)
                         #print(data[i])
                           #print(i)
                           #print(Mari)
                    return Juanpi, Mari
                Kari =[]
                Juanjito=[]
                linear_search(juanjose,target, Juanjito,Kari)    
                #print(len(Juanjito))
                #print(Juanjito)
                #print(len(Kari))
                #print(Juanjito)
                Juanpablo =[]
                Mayita =[]
                
                for i in range(len(Juanjito)):
                    Juanpablo.append(juanjose[Juanjito[i]])
                    Mayita.append(juanjose[Kari[i]])
                mylict5 = (list(zip(Juanpablo,Mayita)))
                #print(mylict5)
                mylict6 = list(zip(size_total,mylict2,mylict3,mylict5))
                print(mylict6)
            
    def justamethod (self,event):
            print("method is called")
            print (self.cbox.get())
                   
    def pixels(self):
    
        import matplotlib
        from matplotlib import pyplot as plt
        matplotlib.use("TkAgg")
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2TkAgg
        from matplotlib.figure import Figure
        import matplotlib.patches as mpatches
        
        def _quit():
            master.quit()     # stops mainloop
            master.destroy()  # this is necessary on Windows to prevent
            #self.root2.deiconify()                # Fatal Python Error: PyEval_RestoreThread: NULL tstate

        # LARGE_FONT= ("Verdana", 12)
        
        master = tk.Tk()
        master.wm_title("Pixels en la imagen: Espacio Red, Green, Blue. ")
        master.tkraise()
        #self.root2.withdraw() #.tkraise()

      
        color = '#fee2a1'
        self.canvas = Canvas(self.canvas, width=3, height=3)
        self.canvas.config(bg = color)
        self.canvas.pack()
        #self.button = tk.Button(master, text='Quit', command=_quit)
        #self.button.pack() #side=tk.Tk.BOTTOM)
            
        image = cv2.imread("D://Users//jtaquia//.spyder-py3//escena2.jpg")
        
        chans = cv2.split(image)
        colors = ("b", "g", "r")
        self.f = plt.figure()
        #red_patch = mpatches.Patch(color='red', label='Histograma RGB')
        #self.f.legend(handles=[red_patch])
        
        #self.f.legend("Histograma RGB")
        #self.f..xlabel("Bins")
        #self.f.ylabel("Pixels")
       
            
        self.f = Figure(figsize=(3,3), dpi=150)
        self.a = self.f.add_subplot(111)

        for (chan, color) in zip(chans, colors):
            hist = cv2.calcHist([chan], [0], None, [250], [0, 250])
            self.a.plot(hist, color = color)
            self.a.plot([0, 256])
            self.a.plot([0, 15256])
    

        self.canvas = FigureCanvasTkAgg(self.f,master)
        self.canvas.show()

        self.canvas.get_tk_widget().pack() #side=root2.BOTTOM, fill=root2.BOTH, expand=True)

        self.toolbar = NavigationToolbar2TkAgg(self.canvas,master)
        self.toolbar.update()
        self.canvas._tkcanvas.pack() #side=root2.TOP, fill=root2.BOTH, expand=True)
          
    def Reinicio(self):
            self.Text.delete('1.0', END)
    
    def mostrar500(self):
            import os
            import re
            
            w = self.lstbox #.widgetName(Listbox(self.root2))
            index = int(w.curselection()[0])
            value = w.get(index)
            print ('You selected item %d: "%s"' % (index, value))
            nombre = re.split('\:+',value)    
            print(nombre)        
            
            #lb = Listbox(frame, name='lb')
            self.lstbox.bind('<<ListboxSelect>>') #, onselect(value))
            
         
            s = value
            #print('El valor de s es : ' + s)
            start = s.find('página') + 7
                          
            #print('DIVISION ENTRE START y END: ' + str(start))
            end = s.find('esta', start)
            r = s[start:end]              
            
            self.words = r.split()
            #print('EL  r.split() de self.words es : ' + "".join((s+ ' ') for s in self.words))
            
            if   len(self.words)>1 :
                        
                messagebox.showwarning("Warning","Hay varias páginas en la selección elegida. Ingrese el número de página a analizar. " )    
              
                self.root4 = tk.Tk()
                self.root4.title("Análisis de imagen en catálogo.")
                self.root4.geometry("350x150")
                self.label22 = tk.Label(self.root4, text="Ingrese página:", width = 20, height=2, font=("Helvetica", 12))
                self.label22.place(x =5 , y =20)
                self.root4.attributes("-topmost", True)
                color = '#f6b06f'
                self.root4.config(bg = color)
                var_jat = StringVar()
                self.e = Entry( self.root4, textvariable=var_jat)
                self.e.place(x=200, y = 40)
                #button22 = tk.Button(root4, text='Continuar',width = 10,command =  Activa.__init__, height=2, font=("Helvetica", 12))
                self.button22 = tk.Button(self.root4, text='Continuar',width = 10,command =  self.muestra_unica, height=2, font=("Helvetica", 12))
                self.button22.place(x =40 , y =80)
                self.button23 = tk.Button(self.root4, text='Salir',width = 10,command =  self.ask_quit, height=2, font=("Helvetica", 12))
                self.button23.place(x =160 , y =80)
                
                
                #inicio = e.get()
                    #path_to_pdf_file = os.path.abspath(os.path.join( self.claves, str(self.words_file) ))    
                    
                #txtPathfile = os.path.abspath(os.getcwd() + '\document-page'+ str(inicio) + '.pdf')
                 
                
                #new_path2 = os.getcwd() + '//document-page' + inicio + '.pdf'
                #print('el objeto txtPathfile es :' + txtPathfile)
                #root4.destroy()
                #self.root2.deiconify()
                #from wand.image import Image
         
                #with Image(filename= txtPathfile) as img:
                #    print('pages = ', len(img.sequence))
                #    with img.convert('png') as converted:
                #                converted.save(filename='document-page'+ str(inicio) + '.png')
                
            else:
                
                
                inicio = self.words[0]
                print('el valor de la variable inicio es : ' + str(inicio))
                txtPathfile = os.path.abspath(os.getcwd() + '\document-page'+ str(inicio) + '.pdf')
                #txtPathfile = 'D://Yanbal_DLO//PROYECTO COMPUTER VISION//CATALOGOS_PDF//Analisis6//document-page110.pdf'
                #txtPathfile = txtPathfile.replace('/','//')
                #print(txtPathfile)
                from wand.image import Image
                with Image(filename= txtPathfile) as img:
                    print('pages = ', len(img.sequence))
                    with img.convert('png') as converted:
                                converted.save(filename='document-page'+ str(inicio) + '.png')
                #numero_page = int(str(variable.get()))
                
            # luego de aplica script con WAND utilizando ese pdf para extraer (convertir) la imagen con extencion PNG
            
            # y le aplica el script de histograma RGB para mostrar la imagen y el histograma respectivo
            # la idea de esta parte es tener con esa imagen PNG diversos scripts que salen del boton del menu para
            # determinar cuanto vendio esa pagina, hacer canny detection, 
            # vincular ese histograma con otros histogramas similares en el pasado
    
    def mostrar100(self):
            import os
            import re
         
           # def onselect(evt,value):
                # Note here that Tkinter passes an event object to onselect()
                #w = evt.widget
            w = self.lstbox #.widgetName(Listbox(self.root2))
            index = int(w.curselection()[0])
            value = w.get(index)
            #print ('You selected item %d: "%s"' % (index, value))
                
            
            #lb = Listbox(frame, name='lb')
            self.lstbox.bind('<<ListboxSelect>>') #, onselect(value))
            
         
            s = value
            #print('El valor de s es : ' + s)
            start = s.find('página :') + 8
            end = s.find('esta', start)
            r = s[start:end]              
            
            self.words = r.split()
            #words = r.split(",") 
            #print('El valor de words es : ' +  " ".join([str(x) for x in self.words] ))
            
            self.sfile = value
            #print('El valor de sfile es : ' + self.sfile)
            start_sfile = self.sfile.find('archivo:') + 9
            end_sfile = self.sfile.find(':', start_sfile)
            global r_file
            r_file = self.sfile[start_sfile:end_sfile]            
            r_file = r_file.strip()
            self.words_file = r_file #.split()
            
            
            Jjose = 'AcroRd32.exe'
            for root, dirs, files in os.walk(r'C://Program Files (x86)//Adobe//'):
                                       for name in files:
                                           if name == Jjose:
                                               path1 = os.path.abspath(os.path.join(root, name))
                                               
                                   
            path_to_acrobat = os.path.abspath(path1) 
           
            #print(path_to_acrobat)
            
            from subprocess import check_output,Popen, PIPE
                        
            for m in self.words:
               
                path_to_pdf = os.path.abspath(os.path.join( self.claves, str(self.words_file) ))    
                txtPathPdf = path_to_pdf.replace('/','//')
                #print (txtPathPdf)
                
                try:
                    lsout=Popen(['lsof',txtPathPdf],stdout=PIPE, shell=False)
                    check_output(["grep",txtPathPdf], stdin=lsout.stdout, shell=False)
                except:
               #check_output will throw an exception here if it won't find any process using that file
                
                    process = subprocess.Popen([path_to_acrobat, '/A', 'page=' + str(m), path_to_pdf], shell=False, stdout=subprocess.PIPE)
            
    
    def removethis2(self):
        
            #self.root2.quit()
            self.root2.destroy()
            root.deiconify()
    
      
        
    def ask_quit(self):
            
        import tkinter.messagebox
            
        if tkinter.messagebox.askokcancel("Culminar esta búsqueda", "¿Desea salir totalmente de estos resultados? "):
                self.root4.destroy()
            
        self.root2.deiconify()
        print("se termino busqueda")
            
            
    def ejecuta(self): 
            
            
       
            from PyPDF2 import PdfFileWriter, PdfFileReader
            from tkinter import filedialog
            import os
         
          
            w = self.lstbox #.widgetName(Listbox(self.root2))
            index = int(w.curselection()[0])
            value = w.get(index)
            #print ('You selected item %d: "%s"' % (index, value))
                
            
            #lb = Listbox(frame, name='lb')
            self.lstbox.bind('<<ListboxSelect>>') #, onselect(value))
            
         
            s = value
            #print('El valor de s es : ' + s)
            start = s.find('página :') + 8
            end = s.find('esta', start)
            r = s[start:end]              
            
            self.words = r.split()
            #words = r.split(",") 
            #print('El valor de words es : ' +  " ".join([str(x) for x in self.words] ))
            #print('el valor de len: self.words : ' + str(len(self.words)))
            self.sfile = value
            #print('El valor de sfile es : ' + self.sfile)
            start_sfile = self.sfile.find('archivo:') + 9
            end_sfile = self.sfile.find(':', start_sfile)
            global r_file
            r_file = self.sfile[start_sfile:end_sfile]            
            r_file = r_file.strip()
            self.words_file = r_file #.split()
    
            messagebox.showwarning("Warning","Va a realizar un análisis de composición de imágenes y su impacto en la venta. " )
            
            currdir = os.getcwd() 
            
            tempdir3 = filedialog.askdirectory( initialdir= currdir, title='Seleccione donde guardará el resultado')
            
            txtDir100 = tempdir3.replace('/','//')+ "//" 
            txtDir = tempdir3.replace('/','//')+ "//" + r_file
            #print(txtDir)
            path_mari = os.path.abspath(txtDir) 
            #print(path_mari)
            inputpdf = PdfFileReader(open( path_mari, "rb"))
         
            newpath = txtDir100 + 'Analisis8' 
            if not os.path.exists(newpath):
                os.makedirs(newpath)
    
            os.chdir(newpath)
            
            
            for i in range(0, inputpdf.getNumPages()):
                output = PdfFileWriter()
                output.addPage(inputpdf.getPage(i))
                with open("document-page%s.pdf" % i, "wb") as outputStream:
                    output.write(outputStream)
                    
            
            # encuentra la pagina en pdf de la que se debe extraer la imagen utilizando walk
            if  len(self.words)>1 :
                        
                messagebox.showwarning("Warning","Hay varias páginas en la selección elegida. Ingrese el número de página a analizar. " )    
                self.root4 = Tk()
                self.root4.title("Análisis de imagen en catálogo.")
                self.root4.geometry("350x150")
                self.label22 = tk.Label(root4, text="Ingrese página:", width = 20, height=2, font=("Helvetica", 12))
                self.label22.place(x =5 , y =20)
                self.root4.attributes("-topmost", True)
                color = '#f6b06f'
                self.root4.config(bg = color)
                var_jat = StringVar()
                self.e = Entry( root4, textvariable=var_jat)
                self.e.place(x=200, y = 40)
                #button22 = tk.Button(root4, text='Continuar',width = 10,command =  Activa.__init__, height=2, font=("Helvetica", 12))
                self.button22 = tk.Button(root4, text='Continuar',width = 10,command =  self.muestra_unica, height=2, font=("Helvetica", 12))
                self.button22.place(x =40 , y =80)
              
                
                #inicio = e.get()
                    #path_to_pdf_file = os.path.abspath(os.path.join( self.claves, str(self.words_file) ))    
                    
                #txtPathfile = os.path.abspath(os.getcwd() + '\document-page'+ str(inicio) + '.pdf')
                 
                
                #new_path2 = os.getcwd() + '//document-page' + inicio + '.pdf'
                #print('el objeto txtPathfile es :' + txtPathfile)
                #root4.destroy()
                #self.root2.deiconify()
                #from wand.image import Image
         
                #with Image(filename= txtPathfile) as img:
                #    print('pages = ', len(img.sequence))
                #    with img.convert('png') as converted:
                #                converted.save(filename='document-page'+ str(inicio) + '.png')
                
            else:
                
                
                inicio = self.words[0]
                #print('el valor de la variable inicio es : ' + str(inicio))
                txtPathfile = os.path.abspath(os.getcwd() + '\document-page'+ str(inicio) + '.pdf')
                #txtPathfile = 'D://Yanbal_DLO//PROYECTO COMPUTER VISION//CATALOGOS_PDF//Analisis6//document-page110.pdf'
                #txtPathfile = txtPathfile.replace('/','//')
                #print(txtPathfile)
                from wand.image import Image
                with Image(filename= txtPathfile) as img:
                    print('pages = ', len(img.sequence))
                    with img.convert('png') as converted:
                                converted.save(filename='document-page'+ str(inicio) + '.png')
                #numero_page = int(str(variable.get()))
                
            # luego de aplica script con WAND utilizando ese pdf para extraer (convertir) la imagen con extencion PNG
            
            # y le aplica el script de histograma RGB para mostrar la imagen y el histograma respectivo
            # la idea de esta parte es tener con esa imagen PNG diversos scripts que salen del boton del menu para
            # determinar cuanto vendio esa pagina, hacer canny detection, 
            # vincular ese histograma con otros histogramas similares en el pasado
    
    def muestra_unica(self):
           
             Jjose = 'AcroRd32.exe'
             for root, dirs, files in os.walk(r'C://Program Files (x86)//Adobe//'):
                                       for name in files:
                                           if name == Jjose:
                                               path1 = os.path.abspath(os.path.join(root, name))
                                               
                                   
             path_to_acrobat = os.path.abspath(path1) 
           
             #print(path_to_acrobat)
             index = int(self.lstbox.curselection()[0])
             value = self.lstbox.get(index)
             self.sfile = value
             
             
             start_sfile = self.sfile.find('archivo') + 8
                                          
                                          
             end_sfile = self.sfile.find(':', start_sfile)
             
             
             
             
             global r_file
             r_file = self.sfile[start_sfile:end_sfile]            
             r_file = r_file.strip()
             self.words_file = r_file #.split()
             
             print(self.words_file)
             
             print(str(self.e.get().encode('utf-8')))
             
            
             
             from subprocess import check_output,Popen, PIPE
                        
          
             txtPathfile = os.path.abspath(os.getcwd()) #+ "//" + str(self.words_file) )
             
            # path_to_pdf = os.path.abspath(os.path.join( self.claves, str(self.words_file) ))    
             #txtPathPdf = txtPathfile.replace('/','//')
             path_to_pdf =  txtPathfile + '/'  + str(self.words_file)
             #txtPathPdf = txtPathfile.replace('/','//')
             #print (txtPathPdf)
             print(path_to_pdf)   
             process = subprocess.Popen([path_to_acrobat, '/A', 'page=' + self.e.get(), path_to_pdf], shell=False, stdout=subprocess.PIPE)
              
            #####################################################################
             try:
                    lsout=Popen(['lsof',txtPathPdf],stdout=PIPE, shell=False)
                    check_output(["grep",txtPathPdf], stdin=lsout.stdout, shell=False)
             except:
               #check_output will throw an exception here if it won't find any process using that file
                
                    process = subprocess.Popen([path_to_acrobat, '/A', 'page=' + self.e.get(), path_to_pdf], shell=False, stdout=subprocess.PIPE)
            
             
            
             
            
    def printing1(self):
            import os
            import PyPDF2
            import re
            
            currdir = os.getcwd()
            
            tempdir2 = filedialog.askdirectory( initialdir= currdir, title='Favor seleccion donde guardará el resultado')   
            self.claves = tempdir2
            os.chdir(tempdir2)
            
            k= glob.glob('*.txt')
            
           
            #size_total=[]
            trazos=[]
          

            for i in k:
                
                txtDir0902 = tempdir2.replace('/','//')+ "//" + i 
                
                #size_total.append(size)
                #trazos_total.append(trazos)
                
                import mmap
                fp = open(txtDir0902,'rb')
               
                with fp as file, \
                 mmap.mmap(file.fileno(), 0, access=mmap.ACCESS_READ) as s:
                 v = self.Entry.get().encode('utf-8')
                 juanjito = self.cbox.get() #.encode('utf-8')
                 print(juanjito)
                 trazos = re.split('\s+|\.+',juanjito)
                 print(trazos)
                 v = self.cbox.get().encode('utf-8')
                 
                 if s.find(v) != -1:

                        buffer = io.StringIO()
                        
                        print('La palabra: ' + self.Entry.get() + ' ,si esta en el archivo: ' + i, file=buffer)
                        
                        
                        output = buffer.getvalue()
                        
                       
                        self.Text.insert(END, output)
                         
                
                 else:
                    
                        
                        buffer = io.StringIO()
                        
                        print('La palabra: ' +  self.Entry.get() +  ' ,no se encuentra en el archivo: ' + i, file=buffer)
                        
                        
                        output = buffer.getvalue()
                        
                       
                        self.Text.insert(END, output)
                        
            self.myd = myDict()
            
            Dict_pickin = {}
            os.chdir(tempdir2)
            t= glob.glob('*.pdf')
            
          
                               
            for j in t:
                            Lista_pickin = []
                            pdfDir0902 = tempdir2.replace('/','//')+ "//" + j  
                            #print(pdfDir0902)
                             
                            PageFound = -1
                            pdfDoc = PyPDF2.PdfFileReader(open(pdfDir0902, "rb"))
                            counter =0
                            for i in range(0, pdfDoc.getNumPages()):
                                counter += 1
                                #print(pdfDoc.getNumPages())r
                                content = ""
                                content += pdfDoc.getPage(i).extractText()+ "\n"
                                #print(content)
                                #content1 = content.encode('utf-8') #, 'ignore').lower()
                                #content1 = content.encode('ascii', 'ignore').lower()
                                ResSearch = re.search(self.Entry.get(),content)
                               
                                
                                if ResSearch is not None:
                                   PageFound = i
                                   #print(content)
                                   #var101 = i+1
                                   var101 = i
                                   Lista_pickin.append(var101)
                                   Dict_pickin.update({ str(j):var101}) # default_data.items() + {'item3': 3}.items()) 
                                   
                                   part1 = 'En el archivo '+ j
                                   #part2 = 'En la página : '+ str(var101)
                                   part2 = 'En la página '+ " ".join([str(x)  for x in Lista_pickin] ) #+ str(len(Lista_pickin))
                                   #print(part2)
                                   
                                   self.myd.add(part1 , part2)
                                   
                                   
                                   #Resumen1 = dict(zip(t,counter))                
                                   print(' el resultado SI esta en la página: '+ str(PageFound+1))
                                   path_to_pdf = os.path.abspath( pdfDir0902 )
                                   
                                   Jjose = 'AcroRd32.exe'
                                   for root, dirs, files in os.walk(r'C://Program Files (x86)//Adobe//'):
                                       for name in files:
                                           if name == Jjose:
                                               path1 = os.path.abspath(os.path.join(root, name))
                                               
                                   
                                   path_to_acrobat = os.path.abspath(path1) 
                                   var = PageFound + 1
                                   #process = subprocess.Popen([path_to_acrobat, '/A', 'page=' + str(var), path_to_pdf], shell=False, stdout=subprocess.PIPE)
                                   #process.wait()
                                
                                else:
                                 
                                     print('No se encontró en el archivo : ' + pdfDir0902 + ' en la página ' + str(counter) + ' la palabra que ingreso.')
                                #print(myd)                 
                                    
                                   #for key in sorted(myd):
                                       
           
            for key in sorted(self.myd):    
                self.lstbox.insert(END, '{}: {}'.format(key, self.myd[key])+ ' esta la palabra '+ self.Entry.get())
            
            #print('el resultado de myd es: ' + str(myd))
            messagebox.showwarning("Warning","Se culminó con la búsqeda en los pdf colocados en la carpeta. " )
            
            #print(Dict_pickin)
            #print(myd)
            #print(Lista_pickin)


    def printing2(self):
                import os
                import PyPDF2
                import re
                
                currdir = os.getcwd()
                
                tempdir2 = filedialog.askdirectory( initialdir= currdir, title='Favor seleccion donde guardará el resultado')   
                self.claves = tempdir2
                os.chdir(tempdir2)
                
                k= glob.glob('*.txt')
                
                size=10
                trazos = 100
                
                size_total=[]
                trazos_total=[]
              
    
                for i in k:
                    
                    txtDir0902 = tempdir2.replace('/','//')+ "//" + i 
                    
                    size_total.append(size)
                    trazos_total.append(trazos)
                    import mmap
                
                
                    fp = open(txtDir0902,'rb')
                   
                    with fp as file, \
                     mmap.mmap(file.fileno(), 0, access=mmap.ACCESS_READ) as s:
                     v = self.Entry.get().encode('utf-8')
                     
                     if s.find(v) != -1:
                            
                            
                            
                            buffer = io.StringIO()
                            
                            print('La palabra: ' + self.Entry.get() + ' ,si esta en el archivo: ' + i, file=buffer)
                            
                            
                            output = buffer.getvalue()
                            
                           
                            self.Text.insert(END, output)
                             
                    
                     else:
                        
                            
                            buffer = io.StringIO()
                            
                            print('La palabra: '
                                  
                                  +  self.Entry.get() +  ' ,no se encuentra en el archivo: ' + i, file=buffer)
                            
                            
                            output = buffer.getvalue()
                            
                           
                            self.Text.insert(END, output)
                            
                myd = myDict()
                
                Dict_pickin = {}
                os.chdir(tempdir2)
                t= glob.glob('*.pdf')
                
              
                                   
                for j in t:
                                Lista_pickin = []
                                pdfDir0902 = tempdir2.replace('/','//')+ "//" + j  
                                #print(pdfDir0902)
                                 
                                PageFound = -1
                                pdfDoc = PyPDF2.PdfFileReader(open(pdfDir0902, "rb"))
                                counter =0
                                for i in range(0, pdfDoc.getNumPages()):
                                    counter += 1
                                    #print(pdfDoc.getNumPages())r
                                    content = ""
                                    content += pdfDoc.getPage(i).extractText() + "\n"
                                    #print(content)
                                    content1 = content.encode('utf-8') #, 'ignore').lower()
                                    #content1 = content.encode('ascii', 'ignore').lower()
                                    ResSearch = re.search(self.Entry.get().encode('utf-8') , content1)
                                    
                                    if ResSearch is not None:
                                       PageFound = i
                                       #var101 = i+1
                                       var101 = i
                                       Lista_pickin.append(var101)
                                       Dict_pickin.update({ str(j):var101}) # default_data.items() + {'item3': 3}.items()) 
                                       
                                       part1 = 'En el archivo: '+ j
                                       #part2 = 'En la página : '+ str(var101)
                                       part2 = 'En la página : '+ " ".join([str(x) for x in Lista_pickin] ) #+ str(len(Lista_pickin))
                                       #print(part2)
                                       myd.add(part1 , part2)
                                       #Resumen1 = dict(zip(t,counter))                
                                       print(' el resultado SI esta en la página: '+ str(PageFound))
                                       path_to_pdf = os.path.abspath( pdfDir0902 )
                                       
                                       Jjose = 'AcroRd32.exe'
                                       for root, dirs, files in os.walk(r'C://Program Files (x86)//Adobe//'):
                                           for name in files:
                                               if name == Jjose:
                                                   path1 = os.path.abspath(os.path.join(root, name))
                                                   
                                       
                                       path_to_acrobat = os.path.abspath(path1) 
                                       var = PageFound + 1
                                       #process = subprocess.Popen([path_to_acrobat, '/A', 'page=' + str(var), path_to_pdf], shell=False, stdout=subprocess.PIPE)
                                       #process.wait()
                                    
                                    else:
                                     
                                         print('No se encontró en el archivo : ' + pdfDir0902 + ' en la página ' + str(counter) + ' la palabra que ingreso.')
                                    #print(myd)                 
                                        
                                       #for key in sorted(myd):
                for key in sorted(myd):    
                    self.lstbox.insert(END, '{}: {}'.format(key, myd[key])+ ' esta la palabra '+ self.Entry.get())
                
                messagebox.showwarning("Warning","Se culminó con la búsqeda en los pdf colocados en la carpeta. " )
                #print(Dict_pickin)
                #print(myd)
                #print(Lista_pickin)


class Activa():
            
             

        def _init_(self):
            #print(" el valor de la hoja elegida es :" + e.get())
            global inicio 
            inicio = GUI2.ejecuta.e.get()
            print(inicio.decode('utf-8'))
            #path_to_pdf_file = os.path.abspath(os.path.join( self.claves, str(self.words_file) ))    
            
            new_path2 = os.path.abspath( os.getcwd() + '//document-page'+ inicio + '.pdf')    
            txtPathfile = new_path2.replace('/','//')
            #new_path2 = os.getcwd() + '//document-page' + inicio + '.pdf'
            print(txtPathfile)
            
            self.root2.deiconify()
            from wand.image import Image
     
            with Image(filename= txtPathfile) as img:
                print('pages = ', len(img.sequence))
                with img.convert('png') as converted:
                            converted.save(filename='document-page'+ str(inicio) + '.png')
            
            return txtPathfile         
                
class GUI1:

        def __init__(self, menubar):
            
            from PIL import ImageTk, Image

            tempdir = "jat"
            tempdir2 = "kari"
            image = Image.open('D://Documentos//JOSE ANTONIO//COMPUTER VISION//EJECUTABLE//gif1.gif')
            image2 = Image.open('D://Documentos//JOSE ANTONIO//COMPUTER VISION//EJECUTABLE//gif3.gif')
            photo = ImageTk.PhotoImage(image)
            photo2 = ImageTk.PhotoImage(image2)
            self.label = tk.Label(image=photo)
            self.label.image = photo # keep a reference!
            self.label.place(x=210,y=90)
            self.label2 = tk.Label(image=photo2)
            self.label2.image2 = photo2 # keep a reference!
            self.label2.place(x=400,y=90)
            self.label3 = tk.Label(root,text="Fase 1: Analítica con imágenes:", width = 30, height=4, fg = "red", bg = "white",font=("Helvetica",12,"bold"))
            self.label3.place(x=180, y  = 2)
                    
                    
            filemenu = tk.Menu(menubar, tearoff=0)
            filemenu.add_command(label="Procesar pdfs", command=lambda: ClassA.abrir1(tempdir,tempdir2))

            filemenu.add_command(label="Guardar resultados", command=hello)
            filemenu.add_command(label="Búsqueda", command= entry_fields)
            
            filemenu.add_separator()
            filemenu.add_command(label="Salir", command= salida)
            menubar.add_cascade(label="File", menu=filemenu)
            
            # create more pulldown menus
            editmenu = tk.Menu(menubar, tearoff=0)
            editmenu.add_command(label="Cut", command=hello)
            editmenu.add_command(label="Copy", command=hello)
            editmenu.add_command(label="Paste", command=hello)
            menubar.add_cascade(label="Edit", menu=editmenu)
            
            helpmenu = tk.Menu(menubar, tearoff=0)
            helpmenu.add_command(label="About", command=hello)
            helpmenu.add_command(label="Info", command=Informacion)
            menubar.add_cascade(label="Help", menu=helpmenu)
            #color = '#A3A5DE'
            color = '#fee2a1'
            root.config(menu=menubar, bg=color)
            root.title("Aplicación que busca texto en múltiples PDF´s")
            root.geometry("700x400")
            root.attributes("-topmost", True)
            
        def removethis(self):
            self.frame.destroy()

def Informacion():
        
            #root.withdraw()
            root3 = Tk()
            root3.title("Información")
            root3.geometry("800x100")
            root3.attributes("-topmost", True)
            #color = '#F57723'
            #root3.config(bg = 'white')
            color = '#fee2a1'
            root3.config(bg = color)
            #root2.tkraise()
            window = GUI3(root3)


            
def entry_fields():
        
            root.withdraw()
            root2 = tk.Tk()
            #root4 = tk.Tk()
            root2.title("Búsqueda")
            root2.geometry("650x550")
            root2.attributes("-topmost", True)
            color = '#f6b06f'
            #color = '#F57723'
            root2.config(bg = color)
            #root2.tkraise()
            menu = tk.Menu(root2) #, tearoff=0)
            #menu = tk.Menu(root2) #, tearoff=0)
            root2.config(menu= menu)        
            menu.add_command(label="Espacio RGB", command=setBgColor)
            menu.add_command(label="Composición", command= hello)
            menu.add_command(label="Análisis Precios")
            #menu.add_cascade(label="File", menu=menu)
            window = GUI2(root2)
          

     
def setBgColor():
            (triple, hexstr) = askcolor()
            if hexstr:
                print (hexstr)
                #push.config(bg=hexstr)
     
           
        
def cambios():
        
        root.deiconify()
       # master1.grid_forget()

    
root = Tk()     
menubar = tk.Menu(root)
window = GUI1(menubar)
root.mainloop()







